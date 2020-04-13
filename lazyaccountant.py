#! python3
import re
import pyperclip
import openpyxl
from openpyxl import load_workbook
import os
# open pdf or and copy all to clipboard

# Create a regex for numbers
normalinvoiceRegex = re.compile(r'''
\d{7}
''', re.VERBOSE)
ABCDinvoiceRegex = re.compile(r'''
ABCD\s?\d{3}
''', re.VERBOSE)

# Get the text off the clipboard
text = pyperclip.paste()

# Extract the number from this text
normalextractednumber = normalnumberRegex.findall(text)
ABCDextractednumber = ABCDnumberRegex.findall(text.replace(" ", ""))

allnumbers = []
for Number in normalextractednumber:
    allnumbers.append(Number)
for Number in ABCDextractednumber:
    allnumbers.append(Number)

# Open excel and insert results Column A
os.chdir(r'C:\path')
wb = load_workbook(filename='filename.xlsx')
ws = wb["working sheet name"]
ws['A1'] = ('collected numbers')
for row, i in enumerate(allnumbers):
    column_cell = 'A'
    ws[column_cell+str(row+2)] = str(i)
wb.save('cross reference excel from database.xlsx')

# Create excel and insert extracted numbers to Column A
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = ('collected numbers')
for row, i in enumerate(allnumbers):
    column_cell = 'A'
    ws[column_cell+str(row+2)] = str(i)
wb.save('collected numbers.xlsx')

# fun stuff
print('All done boss!')
